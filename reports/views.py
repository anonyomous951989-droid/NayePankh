"""
==============================================================
NayePankh Foundation - Reports Views
==============================================================
Views for generating reports: volunteer registration, event
participation, attendance, and monthly statistics.
==============================================================
"""

import io
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Q
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import datetime, timedelta

from volunteers.models import VolunteerProfile
from events.models import Event, VolunteerAssignment, Attendance


@login_required
def reports_home(request):
    """Reports dashboard - overview of all available reports."""
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return render(request, 'dashboard/volunteer_dashboard.html')

    return render(request, 'reports/reports_home.html')


@login_required
def volunteer_report(request):
    """
    Volunteer Registration Report with date filters.
    Shows registration trends, status breakdown, and demographic data.
    """
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return render(request, 'dashboard/volunteer_dashboard.html')

    # Date filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    volunteers = VolunteerProfile.objects.select_related('user').all()

    if date_from:
        volunteers = volunteers.filter(created_at__date__gte=date_from)
    if date_to:
        volunteers = volunteers.filter(created_at__date__lte=date_to)

    # --- Stats ---
    total = volunteers.count()
    approved = volunteers.filter(status='approved').count()
    pending = volunteers.filter(status='pending').count()
    rejected = volunteers.filter(status='rejected').count()

    # --- By City ---
    by_city = volunteers.values('city').annotate(count=Count('id')).order_by('-count')[:10]

    # --- By Gender ---
    by_gender = volunteers.values('gender').annotate(count=Count('id'))

    # --- Monthly Trend ---
    monthly_trend = volunteers.annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(count=Count('id')).order_by('month')

    context = {
        'total': total,
        'approved': approved,
        'pending': pending,
        'rejected': rejected,
        'by_city': by_city,
        'by_gender': by_gender,
        'monthly_trend': monthly_trend,
        'volunteers': volunteers[:20],
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'reports/volunteer_report.html', context)


@login_required
def event_report(request):
    """
    Event Participation Report.
    Shows event statistics and volunteer participation.
    """
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return render(request, 'dashboard/volunteer_dashboard.html')

    events = Event.objects.annotate(
        assignment_count=Count('assignments'),
        attendance_present=Count('attendances', filter=Q(attendances__status='present')),
    ).order_by('-date')

    # Date filter
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        events = events.filter(date__gte=date_from)
    if date_to:
        events = events.filter(date__lte=date_to)

    # Overall stats
    total_events = events.count()
    upcoming = events.filter(status='upcoming').count()
    completed = events.filter(status='completed').count()

    context = {
        'events': events,
        'total_events': total_events,
        'upcoming': upcoming,
        'completed': completed,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'reports/event_report.html', context)


@login_required
def attendance_report(request):
    """
    Attendance Report.
    Shows attendance statistics per event and per volunteer.
    """
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return render(request, 'dashboard/volunteer_dashboard.html')

    # Per-event attendance summary
    events_with_attendance = Event.objects.annotate(
        total_assigned=Count('assignments'),
        total_present=Count('attendances', filter=Q(attendances__status='present')),
        total_absent=Count('attendances', filter=Q(attendances__status='absent')),
        total_late=Count('attendances', filter=Q(attendances__status='late')),
    ).filter(status__in=['ongoing', 'completed']).order_by('-date')

    context = {
        'events': events_with_attendance,
    }
    return render(request, 'reports/attendance_report.html', context)


@login_required
def monthly_statistics(request):
    """
    Monthly Statistics Report.
    Shows aggregated monthly data for registrations, events, and attendance.
    """
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return render(request, 'dashboard/volunteer_dashboard.html')

    # Monthly registrations
    monthly_registrations = VolunteerProfile.objects.annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(count=Count('id')).order_by('-month')[:12]

    # Monthly events
    monthly_events = Event.objects.annotate(
        month=TruncMonth('date')
    ).values('month').annotate(count=Count('id')).order_by('-month')[:12]

    # Current month stats
    now = timezone.now()
    current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    current_month = {
        'registrations': VolunteerProfile.objects.filter(created_at__gte=current_month_start).count(),
        'events': Event.objects.filter(date__gte=current_month_start.date()).count(),
        'approvals': VolunteerProfile.objects.filter(
            status='approved',
            updated_at__gte=current_month_start
        ).count(),
    }

    context = {
        'monthly_registrations': monthly_registrations,
        'monthly_events': monthly_events,
        'current_month': current_month,
    }
    return render(request, 'reports/monthly_stats.html', context)


@login_required
def export_report_excel(request, report_type):
    """Export a specific report to Excel."""
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return render(request, 'dashboard/volunteer_dashboard.html')

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='e94560', end_color='e94560', fill_type='solid')

        if report_type == 'attendance':
            ws.title = 'Attendance Report'
            headers = ['Event', 'Date', 'Volunteer', 'Status', 'Check In', 'Check Out']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            attendances = Attendance.objects.select_related('event', 'volunteer__user').all()
            for row, att in enumerate(attendances, 2):
                ws.cell(row=row, column=1, value=att.event.title)
                ws.cell(row=row, column=2, value=str(att.event.date))
                ws.cell(row=row, column=3, value=att.volunteer.full_name)
                ws.cell(row=row, column=4, value=att.get_status_display())
                ws.cell(row=row, column=5, value=str(att.check_in or ''))
                ws.cell(row=row, column=6, value=str(att.check_out or ''))

            filename = 'attendance_report.xlsx'

        elif report_type == 'events':
            ws.title = 'Event Report'
            headers = ['Title', 'Date', 'Location', 'Status', 'Volunteers Assigned', 'Attendance']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill

            events = Event.objects.all()
            for row, event in enumerate(events, 2):
                ws.cell(row=row, column=1, value=event.title)
                ws.cell(row=row, column=2, value=str(event.date))
                ws.cell(row=row, column=3, value=event.location)
                ws.cell(row=row, column=4, value=event.get_status_display())
                ws.cell(row=row, column=5, value=event.assigned_count)
                ws.cell(row=row, column=6, value=event.attendance_count)

            filename = 'event_report.xlsx'
        else:
            messages.error(request, 'Invalid report type.')
            return render(request, 'reports/reports_home.html')

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except ImportError:
        messages.error(request, 'Excel export not available.')
        return render(request, 'reports/reports_home.html')
