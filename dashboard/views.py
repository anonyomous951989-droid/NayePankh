"""
==============================================================
NayePankh Foundation - Dashboard Views
==============================================================
Views for the volunteer dashboard and admin dashboard.
Includes volunteer management, statistics, notifications, and data export.
==============================================================
"""

import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone
from django.views.decorators.http import require_POST

from accounts.models import User
from accounts.utils import send_approval_email, send_rejection_email
from volunteers.models import VolunteerProfile
from events.models import Event, VolunteerAssignment, Attendance, EventInterest, Notification


@login_required
def dashboard_home(request):
    """
    Route users to the correct dashboard based on their role.
    - Admin users → Admin Dashboard
    - Volunteer users → Volunteer Dashboard
    """
    if request.user.is_staff or request.user.is_admin_user:
        return admin_dashboard(request)
    return volunteer_dashboard(request)


def volunteer_dashboard(request):
    """
    Volunteer's personal dashboard showing:
    - Profile summary and application status
    - Unread notifications with badge count
    - Open events with "I'm Interested" button
    - Events the volunteer expressed interest in (with cancel option)
    - Assigned events
    """
    profile = None
    assigned_events = []
    my_interests = []
    open_events = []
    notifications = []
    unread_count = 0

    try:
        profile = request.user.volunteer_profile

        # --- Assigned Events (last 5) ---
        assigned_events = VolunteerAssignment.objects.filter(
            volunteer=profile
        ).select_related('event').order_by('-event__date')[:5]

        # --- Events volunteer expressed interest in (active, not yet assigned) ---
        assigned_event_ids = set(
            VolunteerAssignment.objects.filter(volunteer=profile).values_list('event_id', flat=True)
        )
        my_interests = EventInterest.objects.filter(
            volunteer=profile,
            is_cancelled=False,
        ).exclude(
            event_id__in=assigned_event_ids
        ).select_related('event').order_by('-submitted_at')

        # --- Open events (upcoming/open_for_interest) the volunteer hasn't expressed interest in ---
        interested_event_ids = set(
            EventInterest.objects.filter(
                volunteer=profile, is_cancelled=False
            ).values_list('event_id', flat=True)
        )
        all_excluded = assigned_event_ids | interested_event_ids
        open_events = Event.objects.filter(
            status__in=['upcoming', 'open_for_interest']
        ).exclude(id__in=all_excluded).order_by('date')[:6]

    except VolunteerProfile.DoesNotExist:
        pass

    # --- Notifications (always fetch for authenticated user) ---
    notifications = Notification.objects.filter(
        recipient=request.user
    ).order_by('-created_at')[:10]
    unread_count = Notification.objects.filter(
        recipient=request.user, is_read=False
    ).count()

    context = {
        'profile': profile,
        'assigned_events': assigned_events,
        'my_interests': my_interests,
        'open_events': open_events,
        'notifications': notifications,
        'unread_count': unread_count,
    }
    return render(request, 'dashboard/volunteer_dashboard.html', context)


def admin_dashboard(request):
    """
    Admin dashboard showing:
    - Statistics cards (total, approved, pending, active events)
    - Recent registrations
    - Events with interested volunteer counts
    - Quick action buttons
    """
    # --- Statistics ---
    total_volunteers = VolunteerProfile.objects.count()
    approved_volunteers = VolunteerProfile.objects.filter(status='approved').count()
    pending_applications = VolunteerProfile.objects.filter(status='pending').count()
    rejected_applications = VolunteerProfile.objects.filter(status='rejected').count()
    active_events = Event.objects.filter(status__in=['upcoming', 'open_for_interest', 'assigned']).count()
    total_events = Event.objects.count()

    # --- Recent Registrations (last 5) ---
    recent_registrations = VolunteerProfile.objects.select_related('user').order_by('-created_at')[:5]

    # --- Upcoming / Open Events with interest counts (next 8) ---
    upcoming_events = Event.objects.filter(
        status__in=['upcoming', 'open_for_interest', 'assigned']
    ).annotate(
        interest_count=Count('interests', filter=Q(interests__is_cancelled=False))
    ).order_by('date')[:8]

    context = {
        'total_volunteers': total_volunteers,
        'approved_volunteers': approved_volunteers,
        'pending_applications': pending_applications,
        'rejected_applications': rejected_applications,
        'active_events': active_events,
        'total_events': total_events,
        'recent_registrations': recent_registrations,
        'upcoming_events': upcoming_events,
    }
    return render(request, 'dashboard/admin_dashboard.html', context)


@login_required
def admin_volunteers_list(request):
    """
    View all volunteers with search, filter, and pagination.
    Admin only.
    """
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')

    volunteers = VolunteerProfile.objects.select_related('user').all()

    # --- Search ---
    search_query = request.GET.get('search', '')
    if search_query:
        volunteers = volunteers.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(skills__icontains=search_query) |
            Q(volunteer_id__icontains=search_query)
        )

    # --- Filter by status ---
    status_filter = request.GET.get('status', '')
    if status_filter:
        volunteers = volunteers.filter(status=status_filter)

    # --- Filter by city ---
    city_filter = request.GET.get('city', '')
    if city_filter:
        volunteers = volunteers.filter(city__icontains=city_filter)

    # --- Pagination (10 per page) ---
    paginator = Paginator(volunteers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get unique cities for filter dropdown
    cities = VolunteerProfile.objects.values_list('city', flat=True).distinct().order_by('city')

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'city_filter': city_filter,
        'cities': cities,
    }
    return render(request, 'dashboard/admin_volunteers.html', context)


@login_required
def admin_volunteer_detail(request, pk):
    """View detailed information about a specific volunteer."""
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')

    profile = get_object_or_404(VolunteerProfile.objects.select_related('user'), pk=pk)
    assignments = profile.assignments.select_related('event').all()
    attendances = profile.attendances.select_related('event').all()

    context = {
        'profile': profile,
        'assignments': assignments,
        'attendances': attendances,
    }
    return render(request, 'dashboard/admin_volunteer_detail.html', context)


@login_required
def approve_volunteer(request, pk):
    """Approve a volunteer's application."""
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')

    profile = get_object_or_404(VolunteerProfile, pk=pk)
    profile.status = 'approved'
    profile.generate_volunteer_id()
    profile.save()

    # Send approval email
    send_approval_email(profile.user)

    messages.success(request, f'{profile.full_name} has been approved as a volunteer!')
    return redirect('dashboard:admin_volunteers')


@login_required
def reject_volunteer(request, pk):
    """Reject a volunteer's application."""
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')

    profile = get_object_or_404(VolunteerProfile, pk=pk)
    reason = request.POST.get('reason', '') if request.method == 'POST' else ''

    profile.status = 'rejected'
    profile.rejection_reason = reason
    profile.save()

    # Send rejection email
    send_rejection_email(profile.user, reason)

    messages.warning(request, f'{profile.full_name}\'s application has been rejected.')
    return redirect('dashboard:admin_volunteers')


# ==============================================================
# NOTIFICATION VIEWS
# ==============================================================

@login_required
def notifications_list(request):
    """Full notifications page for the logged-in volunteer."""
    notifications = Notification.objects.filter(
        recipient=request.user
    ).order_by('-created_at')

    # Mark all as read when viewing the full list
    Notification.objects.filter(
        recipient=request.user, is_read=False
    ).update(is_read=True)

    unread_count = 0  # All marked read now

    paginator = Paginator(notifications, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'dashboard/notifications.html', {
        'page_obj': page_obj,
        'unread_count': unread_count,
    })


@login_required
@require_POST
def mark_notification_read(request, pk):
    """
    Mark a single notification as read.
    Supports both AJAX (returns JSON) and standard POST (redirects).
    """
    notification = get_object_or_404(Notification, pk=pk, recipient=request.user)
    notification.is_read = True
    notification.save()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        remaining_unread = Notification.objects.filter(
            recipient=request.user, is_read=False
        ).count()
        return JsonResponse({'status': 'ok', 'unread_count': remaining_unread})

    return redirect(request.META.get('HTTP_REFERER', 'dashboard:home'))


@login_required
@require_POST
def mark_all_notifications_read(request):
    """Mark all of the user's notifications as read."""
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'ok', 'unread_count': 0})

    messages.success(request, 'All notifications marked as read.')
    return redirect('dashboard:notifications')


# ==============================================================
# DATA EXPORT VIEWS
# ==============================================================

@login_required
def export_volunteers_excel(request):
    """Export all volunteer data to an Excel spreadsheet."""
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = 'Volunteers'

        # --- Header styling ---
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='e94560', end_color='e94560', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # --- Headers ---
        headers = [
            'Volunteer ID', 'First Name', 'Last Name', 'Email', 'Phone',
            'Gender', 'Date of Birth', 'City', 'State', 'Education',
            'Skills', 'Interests', 'Availability', 'Status', 'Registered On'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # --- Data ---
        volunteers = VolunteerProfile.objects.select_related('user').all()
        for row, vol in enumerate(volunteers, 2):
            data = [
                vol.volunteer_id or 'N/A',
                vol.user.first_name,
                vol.user.last_name,
                vol.user.email,
                vol.phone,
                vol.get_gender_display(),
                vol.date_of_birth.strftime('%Y-%m-%d') if vol.date_of_birth else '',
                vol.city,
                vol.state,
                vol.education,
                vol.skills,
                vol.interests,
                vol.get_availability_display(),
                vol.get_status_display(),
                vol.created_at.strftime('%Y-%m-%d %H:%M'),
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border

        # --- Auto-fit column widths ---
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        # --- Save to response ---
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="volunteers_export.xlsx"'
        return response

    except ImportError:
        messages.error(request, 'Excel export is not available. Please install openpyxl.')
        return redirect('dashboard:admin_volunteers')


@login_required
def export_volunteers_pdf(request):
    """Export all volunteer data to a PDF document."""
    if not request.user.is_staff and not request.user.is_admin_user:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        # --- Title ---
        title = Paragraph('NayePankh Foundation - Volunteer Report', styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))

        subtitle = Paragraph(
            f'Generated on: {timezone.now().strftime("%Y-%m-%d %H:%M")}',
            styles['Normal']
        )
        elements.append(subtitle)
        elements.append(Spacer(1, 24))

        # --- Table Data ---
        headers = ['ID', 'Name', 'Email', 'Phone', 'City', 'Status']
        data = [headers]

        volunteers = VolunteerProfile.objects.select_related('user').all()
        for vol in volunteers:
            data.append([
                vol.volunteer_id or 'N/A',
                vol.full_name,
                vol.user.email,
                vol.phone,
                vol.city,
                vol.get_status_display(),
            ])

        # --- Create Table ---
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e94560')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))

        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="volunteers_report.pdf"'
        return response

    except ImportError:
        messages.error(request, 'PDF export is not available. Please install reportlab.')
        return redirect('dashboard:admin_volunteers')
